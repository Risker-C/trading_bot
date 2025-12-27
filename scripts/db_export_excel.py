#!/usr/bin/env python3
"""
数据库导出Excel工具

功能：
- 将数据库表导出为Excel文件
- 每个表一个sheet
- 自动格式化和样式
"""

import sqlite3
import sys
import os
from datetime import datetime

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

DB_PATH = config.DB_PATH


def export_to_excel(output_file=None):
    """导出数据库到Excel"""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("❌ 缺少依赖库，请安装：")
        print("   pip install pandas openpyxl")
        return False

    if output_file is None:
        output_file = f"trading_bot_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print(f"📊 开始导出数据库到 Excel...")
    print(f"数据库: {DB_PATH}")
    print(f"输出文件: {output_file}")

    try:
        conn = sqlite3.connect(DB_PATH)

        # 获取所有表名
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]

        print(f"\n找到 {len(tables)} 个表:")
        for table in tables:
            print(f"  - {table}")

        # 创建Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for table in tables:
                print(f"\n导出表: {table}...", end=" ")

                # 读取表数据
                df = pd.read_sql_query(f"SELECT * FROM {table}", conn)

                # 写入Excel
                df.to_excel(writer, sheet_name=table, index=False)

                print(f"✅ ({len(df)} 行)")

        conn.close()

        # 美化Excel
        print("\n美化Excel格式...", end=" ")
        wb = load_workbook(output_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 设置表头样式
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        print("✅")

        print(f"\n✅ 导出成功！")
        print(f"文件位置: {os.path.abspath(output_file)}")

        return True

    except Exception as e:
        print(f"\n❌ 导出失败: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='导出数据库到Excel')
    parser.add_argument('-o', '--output', help='输出文件名', default=None)
    args = parser.parse_args()

    export_to_excel(args.output)
